from zipfile import ZipFile
from openpyxl import load_workbook
from pypdf import PdfReader
import os.path
import csv
from script_os import TMP_DIR


def test_downloading_file(downloading_file):
    assert os.path.isdir(TMP_DIR)
    assert len(os.listdir(TMP_DIR)) > 0


def test_create_zip(create_zip):
    assert os.path.isfile('zzip.zip')


def test_pdf(create_zip):
    with ZipFile('zzip.zip') as zip_file:
        with zip_file.open(f'sample-5.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            text_pdf = ''
            for page in reader.pages:
                text_pdf += f'{page.extract_text()} \n'

        assert text_pdf is not None and text_pdf.strip() != '', f"Не удалось извлечь текст из страницы PDF {f'sample-5.pdf'}"
        assert 'Your Name' in reader.pages[0].extract_text()
        assert text_pdf.split()[0] in text_pdf
        assert len(reader.pages) > 0, f"PDF файл {f'sample-5.pdf'} пуст!"


def test_csv(create_zip):
    with ZipFile('zzip.zip') as zip_file:
        with zip_file.open(f'sample-5.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            text_csv = ''
            for line in csvreader:
                text_csv += ', '.join(line) + '\n'

        assert ', '.join(csvreader[0]) in text_csv
        assert ', '.join(csvreader[1]) == ', GET STARTED, , , , , , , NOTE, , , '
        assert text_csv is not None, f"CSV файл {f'sample-5.csv'} пуст!"
        assert any(csvreader)


def test_xlsx(create_zip):
    with ZipFile('zzip.zip') as zip_file:
        with zip_file.open(f'sample-5.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            text_xlsx = ''
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text_xlsx += f'{cell.value} \n'

        assert sheet.cell(row=2, column=2).value == 'JANUARY'
        assert text_xlsx.split()[0] in text_xlsx
        assert text_xlsx is not None, f"XLSX файл {f'sample-5.xlsx'} пуст!"
